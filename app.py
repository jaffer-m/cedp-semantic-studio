"""Column Description Review — Streamlit app."""

import io
import os
import csv
from datetime import datetime

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
import ai_gen
import humanize as hz
from databricks_client import DatabricksClient
from catalog import list_tables_via_sql

st.set_page_config(
    page_title="CEDP Semantic Studio",
    page_icon="📖",
    layout="wide",
)

# ── Custom styling ────────────────────────────────────────────────────────

st.markdown("""
<style>
  /* ── Table type badges ─────────────────────────────────── */
  .type-badge {
    display: inline-block; padding: 2px 10px; border-radius: 12px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.5px;
    text-transform: uppercase; margin-left: 10px; vertical-align: middle;
  }
  .badge-dlt   { background: #14532d; color: #4ade80; }
  .badge-mv    { background: #451a03; color: #fbbf24; }
  .badge-delta { background: #1e1b4b; color: #818cf8; }

  /* ── Column name pill ──────────────────────────────────── */
  .col-name {
    display: inline-block; background: #141824; border: 1px solid #2e3755;
    border-radius: 5px; padding: 3px 9px;
    font-family: 'Courier New', monospace; font-size: 13px; color: #c7d2fe;
    word-break: break-all; line-height: 1.5;
  }

  /* ── Data type pills ───────────────────────────────────── */
  .dt-pill {
    display: inline-block; padding: 2px 8px; border-radius: 4px;
    font-size: 11px; font-weight: 600; font-family: monospace; white-space: nowrap;
  }
  .dt-string  { background: #0f2744; color: #7dd3fc; }
  .dt-int     { background: #0a2818; color: #86efac; }
  .dt-double  { background: #132010; color: #a3e635; }
  .dt-bool    { background: #2d1503; color: #fb923c; }
  .dt-date    { background: #1a0f38; color: #c084fc; }
  .dt-other   { background: #161b2e; color: #94a3b8; }

  /* ── Column grid ───────────────────────────────────────── */
  .col-label {
    font-size: 10px; font-weight: 700; letter-spacing: 0.9px;
    text-transform: uppercase; color: #4f5a7a; margin-bottom: 4px;
  }
  .current-desc {
    font-size: 13px; color: #8895b3; padding: 6px 2px;
    min-height: 44px; line-height: 1.55;
  }
  .divider { border-top: 1px solid #1a2035; margin: 6px 0; }

  /* ── Review status chips ───────────────────────────────── */
  .chip {
    display: inline-block; padding: 2px 9px; border-radius: 10px;
    font-size: 11px; font-weight: 600; margin: 3px 2px 0 0; line-height: 1.6;
  }
  .chip-approved { background: #0d3320; color: #4ade80; border: 1px solid #166534; }
  .chip-rejected { background: #2d0a0a; color: #f87171; border: 1px solid #7f1d1d; }
  .chip-note     { background: #111827; color: #94a3b8; border: 1px solid #1e2a3a; }

  /* ── Hint / info card ──────────────────────────────────── */
  .hint-card {
    background: rgba(99,102,241,0.07); border: 1px solid rgba(99,102,241,0.22);
    border-radius: 8px; padding: 11px 16px; font-size: 13px; color: #a5b4fc;
    margin: 4px 0 16px 0; line-height: 1.5;
  }

  /* ── Section headers ───────────────────────────────────── */
  .section-title {
    font-size: 13px; font-weight: 700; color: #cbd5e1;
    margin-bottom: 4px; letter-spacing: 0.2px;
  }
  .section-caption {
    font-size: 12px; color: #4f5a7a; margin-bottom: 10px; line-height: 1.5;
  }

  /* ── Overview label ────────────────────────────────────── */
  .overview-label {
    font-size: 10px; font-weight: 700; letter-spacing: 0.9px;
    text-transform: uppercase; color: #6366f1; margin-bottom: 6px;
  }

  /* ── Empty state ───────────────────────────────────────── */
  .empty-state {
    text-align: center; padding: 64px 32px; color: #4f5a7a;
  }
  .empty-state .es-icon { font-size: 52px; margin-bottom: 18px; line-height: 1; }
  .empty-state h3 { color: #7c8db5; font-size: 20px; margin: 0 0 10px 0; font-weight: 600; }
  .empty-state p  { font-size: 14px; line-height: 1.75; margin: 0; }
  .empty-state .step {
    display: inline-flex; align-items: center; gap: 8px;
    background: #0f1422; border: 1px solid #1e2a3a;
    border-radius: 8px; padding: 8px 16px; margin: 6px 4px;
    font-size: 13px; color: #64748b;
  }
  .empty-state .step-num {
    background: #1e2a45; color: #818cf8; border-radius: 50%;
    width: 20px; height: 20px; display: inline-flex;
    align-items: center; justify-content: center;
    font-size: 11px; font-weight: 700; flex-shrink: 0;
  }
</style>
""", unsafe_allow_html=True)


def _type_pill(dtype: str) -> str:
    dl = dtype.lower()
    if any(x in dl for x in ("string", "varchar", "char", "text")):
        cls = "dt-string"
    elif any(x in dl for x in ("bigint", "int", "long", "short", "byte", "tinyint", "smallint")):
        cls = "dt-int"
    elif any(x in dl for x in ("double", "float", "decimal", "numeric", "real")):
        cls = "dt-double"
    elif "bool" in dl:
        cls = "dt-bool"
    elif any(x in dl for x in ("date", "timestamp")):
        cls = "dt-date"
    else:
        cls = "dt-other"
    return f"<span class='dt-pill {cls}'>{dtype}</span>"


def _type_badge_html(type_label: str) -> str:
    badge_class = (
        "badge-dlt" if "Streaming" in type_label
        else "badge-mv" if "Materialized" in type_label
        else "badge-delta"
    )
    return f"<span class='type-badge {badge_class}'>{type_label}</span>"


# ── Excel export / import helpers ─────────────────────────────────────────

_HEADERS = [
    "table", "column_name", "data_type",
    "current_description", "proposed_description",
    "reviewer_approved",   # reviewer fills: Yes / No
    "reviewer_notes",      # reviewer fills: free text
]

_REVIEWER_COLS = {"reviewer_approved", "reviewer_notes"}
_FILL_HEADER   = PatternFill("solid", fgColor="1E3A5F")
_FILL_REVIEWER = PatternFill("solid", fgColor="FFF9C4")
_FILL_READONLY = PatternFill("solid", fgColor="F5F5F5")
_FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
_FONT_READONLY = Font(color="555555")
_BORDER        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _build_excel(full_name: str, columns: list[dict], suggestions: dict, table_description: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Column Review"

    # Instruction banner (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))
    instr = ws.cell(row=1, column=1,
        value=(
            "Instructions: Fill in 'reviewer_approved' (Yes or No) for each row. "
            "You may also edit 'proposed_description' or add 'reviewer_notes'. "
            "Do not change any other columns. Save and return this file."
        )
    )
    instr.fill = PatternFill("solid", fgColor="E8F4FD")
    instr.font = Font(italic=True, color="1A4A6E", size=10)
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # Table overview row (row 2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_HEADERS))
    td_cell = ws.cell(row=2, column=1,
        value=f"Table Overview: {table_description or '(not generated)'}")
    td_cell.fill = PatternFill("solid", fgColor="E8F0FE")
    td_cell.font = Font(italic=True, color="1A237E", size=10)
    td_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Header row (row 3)
    for col_idx, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[3].height = 28

    # Data rows (row 4+)
    for row_idx, col in enumerate(columns, 4):
        row_data = {
            "table": full_name,
            "column_name": col["name"],
            "data_type": col["type"],
            "current_description": col["current_comment"],
            "proposed_description": suggestions.get(col["name"], ""),
            "reviewer_approved": "",
            "reviewer_notes": "",
        }
        for col_idx, header in enumerate(_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER
            if header in _REVIEWER_COLS:
                cell.fill = _FILL_REVIEWER
            else:
                cell.fill = _FILL_READONLY
                cell.font = _FONT_READONLY

    # Column widths
    for i, w in enumerate([30, 20, 15, 45, 45, 18, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_approved(raw: str) -> bool:
    """Loosely interpret a reviewer's 'approved' cell — tolerates extra words,
    punctuation, or emoji (e.g. 'Yes - looks good', 'Approved ✅') rather than
    requiring an exact 'yes'/'y'/'true'/'1' match."""
    v = raw.strip().lower()
    if not v or v.startswith(("n", "0", "false")):
        return False
    return v.startswith(("y", "true", "1")) or "approve" in v


def _safe_cell(row: tuple, idx: dict, key: str):
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    return row[i]


def _parse_excel(raw: bytes) -> dict[str, dict]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    # Find the header row (contains "column_name")
    header_row_idx = None
    headers = []
    for row in ws.iter_rows():
        vals = [str(c.value or "").strip().lower() for c in row]
        if "column_name" in vals:
            header_row_idx = row[0].row
            headers = vals
            break

    if header_row_idx is None:
        raise ValueError("Could not find a header row with 'column_name'.")

    idx = {h: i for i, h in enumerate(headers)}
    result = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = str(_safe_cell(row, idx, "column_name") or "").strip()
        if not name:
            continue
        approved_raw = str(_safe_cell(row, idx, "reviewer_approved") or "").strip().lower()
        result[name] = {
            "proposed_description": str(_safe_cell(row, idx, "proposed_description") or "").strip(),
            "approved": _is_approved(approved_raw),
            "notes": str(_safe_cell(row, idx, "reviewer_notes") or "").strip(),
        }
    return result


def _parse_csv(raw: bytes) -> dict[str, dict]:
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace")))
    result = {}
    for row in reader:
        name = (row.get("column_name") or "").strip()
        if not name:
            continue
        approved_raw = (row.get("reviewer_approved") or "").strip().lower()
        result[name] = {
            "proposed_description": (row.get("proposed_description") or "").strip(),
            "approved": _is_approved(approved_raw),
            "notes": (row.get("reviewer_notes") or "").strip(),
        }
    return result


# ── Table-level review file helpers (Table only scope) ─────────────────────

_TABLE_HEADERS = [
    "table", "current_description", "proposed_description",
    "reviewer_approved",   # reviewer fills: Yes / No
    "reviewer_notes",      # reviewer fills: free text
]


def _build_table_review_excel(descriptions: dict[str, str], current_comments: dict[str, str]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table Review"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_TABLE_HEADERS))
    instr = ws.cell(row=1, column=1,
        value=(
            "Instructions: Fill in 'reviewer_approved' (Yes or No) for each row. "
            "You may also edit 'proposed_description' or add 'reviewer_notes'. "
            "Do not change any other columns. Save and return this file."
        )
    )
    instr.fill = PatternFill("solid", fgColor="E8F4FD")
    instr.font = Font(italic=True, color="1A4A6E", size=10)
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    for col_idx, header in enumerate(_TABLE_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[2].height = 28

    for row_idx, full_name in enumerate(descriptions, 3):
        row_data = {
            "table": full_name,
            "current_description": current_comments.get(full_name, ""),
            "proposed_description": descriptions.get(full_name, ""),
            "reviewer_approved": "",
            "reviewer_notes": "",
        }
        for col_idx, header in enumerate(_TABLE_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER
            if header in _REVIEWER_COLS:
                cell.fill = _FILL_REVIEWER
            else:
                cell.fill = _FILL_READONLY
                cell.font = _FONT_READONLY

    for i, w in enumerate([40, 50, 50, 18, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_table_review_csv(descriptions: dict[str, str], current_comments: dict[str, str]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_TABLE_HEADERS)
    writer.writeheader()
    for full_name in descriptions:
        writer.writerow({
            "table": full_name,
            "current_description": current_comments.get(full_name, ""),
            "proposed_description": descriptions.get(full_name, ""),
            "reviewer_approved": "",
            "reviewer_notes": "",
        })
    return buf.getvalue().encode("utf-8")


def _parse_table_review_excel(raw: bytes) -> dict[str, dict]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    header_row_idx = None
    headers = []
    for row in ws.iter_rows():
        vals = [str(c.value or "").strip().lower() for c in row]
        if "table" in vals and "proposed_description" in vals:
            header_row_idx = row[0].row
            headers = vals
            break

    if header_row_idx is None:
        raise ValueError("Could not find a header row with 'table' and 'proposed_description'.")

    idx = {h: i for i, h in enumerate(headers)}
    result = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = str(_safe_cell(row, idx, "table") or "").strip()
        if not name:
            continue
        approved_raw = str(_safe_cell(row, idx, "reviewer_approved") or "").strip().lower()
        result[name] = {
            "proposed_description": str(_safe_cell(row, idx, "proposed_description") or "").strip(),
            "approved": _is_approved(approved_raw),
            "notes": str(_safe_cell(row, idx, "reviewer_notes") or "").strip(),
        }
    return result


def _parse_table_review_csv(raw: bytes) -> dict[str, dict]:
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace")))
    result = {}
    for row in reader:
        name = (row.get("table") or "").strip()
        if not name:
            continue
        approved_raw = (row.get("reviewer_approved") or "").strip().lower()
        result[name] = {
            "proposed_description": (row.get("proposed_description") or "").strip(),
            "approved": _is_approved(approved_raw),
            "notes": (row.get("reviewer_notes") or "").strip(),
        }
    return result


# ── Code export helpers ───────────────────────────────────────────────────

def _build_pyspark_script(full_name: str, suggestions: dict, table_description: str = "") -> str:
    quoted = ".".join(f"`{p}`" for p in full_name.split("."))
    lines = [
        f"# Column descriptions for {full_name}",
        f"# Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "from pyspark.sql import SparkSession",
        "spark = SparkSession.builder.getOrCreate()",
        "",
    ]
    if table_description.strip():
        escaped = table_description.replace('"', '\\"')
        lines.append(f'spark.sql("""COMMENT ON TABLE {quoted} IS \\"{escaped}\\"""")')
        lines.append("")
    for col_name, desc in suggestions.items():
        if desc.strip():
            col_safe = col_name.replace("`", "``")
            escaped  = desc.replace('"', '\\"')
            lines.append(
                f'spark.sql("""COMMENT ON COLUMN {quoted}.`{col_safe}` IS \\"{escaped}\\"""")'
            )
    return "\n".join(lines)


def _build_multi_table_pyspark_script(descriptions: dict[str, str]) -> str:
    lines = [
        "# Table-level descriptions",
        f"# Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "from pyspark.sql import SparkSession",
        "spark = SparkSession.builder.getOrCreate()",
        "",
    ]
    for full_name, desc in descriptions.items():
        if desc.strip():
            quoted = ".".join(f"`{p}`" for p in full_name.split("."))
            escaped = desc.replace('"', '\\"')
            lines.append(f'spark.sql("""COMMENT ON TABLE {quoted} IS \\"{escaped}\\"""")')
    return "\n".join(lines)


# ── Session state defaults ────────────────────────────────────────────────

def _init_state():
    defaults = {
        # Connection — host/token pre-seeded from .env if set
        "connected": False,
        "db_client": None,
        "current_user": "",
        "db_host": config.DEFAULT_HOST,
        "db_token": config.DEFAULT_TOKEN,
        "db_endpoint": config.SERVING_ENDPOINT,
        # Browse
        "catalogs": [], "schemas": [], "tables": [], "columns": [],
        "table_load_errors": [],
        "selected_catalog": None, "selected_schemas": [], "selected_table": None,
        "table_meta": {}, "suggestions": {}, "table_description": "",
        "generated": False,
        "gen_error": None,    # persists generation errors across reruns
        "review_import": {},  # {col_name: {approved, notes, proposed_description}}
        # Multi-table ("Table only" scope)
        "table_descriptions": {},  # {full_name: description}
        "table_gen_errors": {},    # {table_name: error string}
        "table_review_import": {},  # {full_name: {approved, notes, proposed_description}}
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()


def _load_tables_for_schemas(
    client: "DatabricksClient", catalog: str, schemas: list[str]
) -> tuple[list[dict], list[str]]:
    """Load and merge tables from one or more schemas, tagging each with a
    schema-qualified name so same-named tables in different schemas don't collide.

    Returns (tables, errors) — a schema with a permission error is skipped
    (not silently dropped) and its message is returned alongside whatever
    other schemas succeeded.
    """
    result = []
    errors = []
    for schema in schemas:
        try:
            tables = client.list_tables(catalog, schema)
        except PermissionError as e:
            errors.append(str(e))
            continue
        if not tables:
            # UC Tables REST API sometimes returns nothing for a schema the
            # SQL warehouse can see fine (platform-side authorization
            # inconsistency) — fall back to information_schema via SQL.
            try:
                tables = list_tables_via_sql(client._ws, catalog, schema)
            except Exception:
                pass
        for t in tables:
            result.append({**t, "schema": schema, "qualified_name": f"{schema}.{t['name']}"})
    return sorted(result, key=lambda x: x["qualified_name"].lower()), errors


# ── Connect helper ────────────────────────────────────────────────────────

def _connect_databricks(host: str, token: str) -> None:
    try:
        client = DatabricksClient(host, token)
        user = client.test_connection()
    except Exception as e:
        st.error(f"Connection failed: {e}")
        return

    st.session_state.db_client = client
    st.session_state.connected = True
    st.session_state.current_user = user
    # Reset browse state so catalogs reload with new credentials
    for key in ("catalogs", "schemas", "tables", "columns", "table_meta",
                "suggestions", "review_import", "table_descriptions",
                "table_gen_errors", "table_review_import", "table_load_errors"):
        st.session_state[key] = [] if isinstance(st.session_state[key], list) else {}
    st.session_state.generated = False
    st.session_state.selected_catalog = None
    st.session_state.selected_schemas = []
    st.session_state.selected_table = None
    st.toast(f"Connected as {user}", icon="✅")


# ── Sidebar ───────────────────────────────────────────────────────────────

def _reset(**extra):
    # Clear text widget keys for the current table so stale values don't bleed
    # into the next table's widgets when column names happen to match.
    for col in st.session_state.get("columns", []):
        st.session_state.pop(f"text_{col['name']}", None)
    st.session_state.pop("text_table_desc", None)
    for full_name in st.session_state.get("table_descriptions", {}):
        st.session_state.pop(f"text_tabledesc_{full_name}", None)
    st.session_state.update(dict(
        columns=[], suggestions={}, table_description="",
        generated=False, gen_error=None, review_import={},
        table_descriptions={}, table_gen_errors={}, ui_tables_multi=[],
        table_review_import={},
        **extra
    ))


with st.sidebar:
    st.title("📖 CEDP Semantic Studio")
    st.caption("Generate and export column descriptions for Unity Catalog tables.")

    # ── Connect form ──────────────────────────────────────────────────────
    connected = st.session_state.connected
    badge = "🟢 Live" if connected else "🔴 Off"
    with st.expander(f"☁️ Databricks Connection  {badge}", expanded=not connected):
        st.text_input(
            "Workspace URL",
            key="db_host",
            placeholder="https://your-workspace.azuredatabricks.net",
        )
        st.text_input(
            "Personal Access Token",
            key="db_token",
            type="password",
            placeholder="dapi…",
        )
        st.text_input(
            "Model Endpoint",
            key="db_endpoint",
            placeholder="databricks-meta-llama-3-3-70b-instruct",
            help="The serving endpoint name in this workspace. Differs between stage and prod.",
        )
        if st.button(
            "Connect",
            type="primary" if not connected else "secondary",
            use_container_width=True,
        ):
            host = st.session_state.db_host.strip()
            token = st.session_state.db_token.strip()
            if not host or not token:
                st.error("Both Workspace URL and token are required.")
            else:
                with st.spinner("Connecting…"):
                    _connect_databricks(host, token)
                st.rerun()

        if connected:
            st.caption(f"Connected as **{st.session_state.current_user}**")

    if not connected:
        st.stop()

    st.divider()

    # ── Catalog browser (only shown after connect) ────────────────────────
    client: DatabricksClient = st.session_state.db_client

    if not st.session_state.catalogs:
        with st.spinner("Loading catalogs…"):
            st.session_state.catalogs = client.list_catalogs()
            if not st.session_state.catalogs:
                st.warning("No catalogs found. Check your USE CATALOG privilege.")

    cat = st.selectbox("Catalog",
        options=["— select —"] + st.session_state.catalogs, key="ui_catalog")

    if cat and cat != "— select —" and cat != st.session_state.selected_catalog:
        _reset(selected_catalog=cat, selected_schemas=[], selected_table=None,
               schemas=[], tables=[], table_meta={}, table_load_errors=[])
        with st.spinner("Loading schemas…"):
            st.session_state.schemas = client.list_schemas(cat)
            if not st.session_state.schemas:
                st.warning(f"No schemas found in `{cat}`. Check your USE SCHEMA privilege.")

    schemas_selected = st.multiselect("Schema(s)",
        options=st.session_state.schemas,
        disabled=not st.session_state.schemas, key="ui_schemas")

    if schemas_selected != st.session_state.selected_schemas:
        _reset(selected_schemas=schemas_selected, selected_table=None, tables=[], table_meta={})
        with st.spinner("Loading tables…"):
            tables, table_errors = _load_tables_for_schemas(client, cat, schemas_selected)
            st.session_state.tables = tables
            st.session_state.table_load_errors = table_errors

    for err in st.session_state.table_load_errors:
        parts = err.split("Raw error:", 1)
        st.error(f"**Access denied** — {parts[0].strip()}")
        if len(parts) > 1:
            with st.expander("Raw error (share with your admin)"):
                st.code(parts[1].strip())
    if (schemas_selected and not st.session_state.tables
            and not st.session_state.table_load_errors):
        st.warning(f"No tables found in the selected schema(s) of `{cat}`.")

    gen_scope = st.radio(
        "Generate scope",
        options=["Table + Columns", "Table only"],
        key="gen_scope",
        horizontal=True,
    )

    if gen_scope == "Table + Columns":
        table_name = st.selectbox("Table",
            options=["— select —"] + [t["qualified_name"] for t in st.session_state.tables],
            disabled=not st.session_state.tables, key="ui_table")

        if table_name and table_name != "— select —":
            tbl_meta = next((t for t in st.session_state.tables if t["qualified_name"] == table_name), None)
            if tbl_meta and tbl_meta["full_name"] != (st.session_state.table_meta or {}).get("full_name"):
                _reset(selected_table=table_name, table_meta=tbl_meta)
                with st.spinner("Loading columns…"):
                    try:
                        st.session_state.columns = client.get_columns(tbl_meta["full_name"])
                    except PermissionError as e:
                        parts = str(e).split("Raw error:", 1)
                        st.error(f"**Access denied** — {parts[0].strip()}")
                        if len(parts) > 1:
                            with st.expander("Raw error (share with your admin)"):
                                st.code(parts[1].strip())
                    except Exception as e:
                        st.error(str(e))
    else:
        sel_col, clr_col = st.columns(2)
        if sel_col.button("Select All", use_container_width=True,
                           disabled=not st.session_state.tables):
            st.session_state.ui_tables_multi = [t["qualified_name"] for t in st.session_state.tables]
            st.rerun()
        if clr_col.button("Clear All", use_container_width=True,
                           disabled=not st.session_state.tables):
            st.session_state.ui_tables_multi = []
            st.rerun()
        st.multiselect(
            "Tables",
            options=[t["qualified_name"] for t in st.session_state.tables],
            disabled=not st.session_state.tables, key="ui_tables_multi",
        )

    st.divider()

    gen_disabled = (
        not st.session_state.columns if gen_scope == "Table + Columns"
        else not st.session_state.ui_tables_multi
    )

    if st.button("⚡ Generate & Humanize",
            disabled=gen_disabled,
            use_container_width=True, type="primary"):
        ws = st.session_state.db_client._ws
        endpoint = st.session_state.db_endpoint.strip() or config.SERVING_ENDPOINT

        if gen_scope == "Table + Columns":
            tbl = st.session_state.table_meta
            cols = st.session_state.columns
            with st.spinner("Generating descriptions…"):
                try:
                    raw = ai_gen.generate_column_descriptions(ws, tbl["full_name"], cols, endpoint)
                except Exception as e:
                    st.session_state.gen_error = str(e)
                    st.rerun()

            st.session_state.gen_error = None
            with st.spinner("Humanizing…"):
                humanized = hz.humanize_all(raw)
            for c in cols:
                val = humanized.get(c["name"], c["current_comment"])
                st.session_state.suggestions[c["name"]] = val
                st.session_state[f"text_{c['name']}"] = val
            with st.spinner("Generating table overview…"):
                try:
                    raw_td = ai_gen.generate_table_description(
                        ws, tbl["full_name"], cols, tbl.get("comment", ""), endpoint
                    )
                    td = hz.humanize(raw_td)
                except Exception:
                    td = st.session_state.table_description
            st.session_state.table_description = td
            st.session_state["text_table_desc"] = td
            st.session_state.review_import = {}

        else:  # Table only
            selected_names = st.session_state.get("ui_tables_multi", [])
            errors = {}
            for name in selected_names:
                tbl_meta = next((t for t in st.session_state.tables if t["name"] == name), None)
                if not tbl_meta:
                    continue
                full_name = tbl_meta["full_name"]
                with st.spinner(f"Generating overview for {name}…"):
                    try:
                        raw_td = ai_gen.generate_table_description(
                            ws, full_name, [], tbl_meta.get("comment", ""), endpoint
                        )
                        td = hz.humanize(raw_td)
                        st.session_state.table_descriptions[full_name] = td
                        st.session_state[f"text_tabledesc_{full_name}"] = td
                    except Exception as e:
                        errors[name] = str(e)
            st.session_state.table_gen_errors = errors
            st.session_state.gen_error = None

        st.session_state.generated = True
        st.rerun()

    if st.session_state.get("gen_error"):
        st.error(st.session_state.gen_error)

    if st.session_state.generated:
        st.caption("✅ Descriptions generated and humanized.")

    st.divider()
    if st.button("⏹ Stop Server", use_container_width=True,
                 help="Shuts down the Streamlit process. Restart with: python -m streamlit run app.py"):
        os._exit(0)


# ── Main area ─────────────────────────────────────────────────────────────

gen_scope = st.session_state.get("gen_scope", "Table + Columns")

if gen_scope == "Table only":
    selected_names = st.session_state.get("ui_tables_multi", [])

    if not selected_names:
        st.markdown("""
<div class='empty-state'>
  <div class='es-icon'>📖</div>
  <h3>Select one or more tables to get started</h3>
  <p>Browse your Unity Catalog from the sidebar, pick tables in the <strong>Tables</strong> multiselect, then generate AI overviews.</p>
</div>
""", unsafe_allow_html=True)
        st.stop()

    st.markdown("### Table-level descriptions")

    for name, err in st.session_state.get("table_gen_errors", {}).items():
        st.error(f"**{name}** — {err}")

    selected_meta = [
        t for t in st.session_state.tables if t["qualified_name"] in selected_names
    ]
    current_comments = {t["full_name"]: t.get("comment", "") for t in selected_meta}
    review_descriptions = {
        t["full_name"]: st.session_state.table_descriptions.get(t["full_name"], t.get("comment", ""))
        for t in selected_meta
    }

    st.divider()
    export_col, import_col = st.columns(2)

    with export_col:
        st.markdown(
            "<div class='section-title'>📤 Share for Review</div>"
            "<div class='section-caption'>Download a review file for your team. "
            "They fill in the approval + notes columns and return it.</div>",
            unsafe_allow_html=True,
        )
        xlsx_col, csv_col = st.columns(2)
        xlsx_col.download_button(
            "⬇ .xlsx",
            data=_build_table_review_excel(review_descriptions, current_comments),
            file_name="table_descriptions_review.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        csv_col.download_button(
            "⬇ .csv",
            data=_build_table_review_csv(review_descriptions, current_comments),
            file_name="table_descriptions_review.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with import_col:
        st.markdown(
            "<div class='section-title'>📥 Import Reviewed File</div>"
            "<div class='section-caption'>Upload the completed file (.xlsx or .csv). "
            "Approved rows load back into the descriptions above.</div>",
            unsafe_allow_html=True,
        )
        uploaded = st.file_uploader(
            "Upload reviewed file", type=["xlsx", "csv"],
            label_visibility="collapsed", key="table_review_upload",
        )
        if uploaded:
            try:
                with st.spinner("Reading reviewed file…"):
                    raw = uploaded.read()
                    review_data = (
                        _parse_table_review_csv(raw) if uploaded.name.endswith(".csv")
                        else _parse_table_review_excel(raw)
                    )
                approved_count = sum(1 for v in review_data.values() if v["approved"])
                rejected_count = len(review_data) - approved_count

                for full_name, data in review_data.items():
                    if data["approved"] and data["proposed_description"]:
                        val = data["proposed_description"]
                        st.session_state.table_descriptions[full_name] = val
                        st.session_state[f"text_tabledesc_{full_name}"] = val

                st.session_state.table_review_import = review_data
                st.success(f"✅ {approved_count} approved  •  ⛔ {rejected_count} not approved")
                st.rerun()
            except Exception as e:
                st.error(f"Could not read file: {e}")

    st.divider()

    export_map: dict[str, str] = {}
    for name in selected_names:
        tbl_meta = next((t for t in st.session_state.tables if t["qualified_name"] == name), None)
        if not tbl_meta:
            continue
        full_name = tbl_meta["full_name"]
        type_label = tbl_meta.get("type_label", "Delta Table")

        review = st.session_state.table_review_import.get(full_name)
        review_html = ""
        if review:
            if review["approved"]:
                review_html += " <span class='chip chip-approved'>✓ Approved</span>"
            else:
                review_html += " <span class='chip chip-rejected'>✗ Rejected</span>"

        st.markdown(
            f"#### `{full_name}`{_type_badge_html(type_label)}{review_html}",
            unsafe_allow_html=True,
        )
        if review and review["notes"]:
            st.markdown(f"<span class='chip chip-note'>💬 {review['notes']}</span>", unsafe_allow_html=True)
        if tbl_meta.get("comment"):
            st.caption(f"Current in Databricks: {tbl_meta['comment']}")

        row_left, row_right = st.columns([6, 1])
        val = row_left.text_area(
            f"tabledesc_{full_name}",
            value=st.session_state.table_descriptions.get(full_name, tbl_meta.get("comment", "")),
            height=80,
            label_visibility="collapsed",
            key=f"text_tabledesc_{full_name}",
            placeholder="Click ⚡ Generate & Humanize to produce an AI overview, or type one manually.",
        )
        st.session_state.table_descriptions[full_name] = val
        export_map[full_name] = val
        with row_right:
            st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
            if st.button("✨", key=f"hz_tabledesc_{full_name}", help=f"Re-humanize {full_name}"):
                with st.spinner("Humanizing…"):
                    humanized_val = hz.humanize(val)
                st.session_state.table_descriptions[full_name] = humanized_val
                st.session_state[f"text_tabledesc_{full_name}"] = humanized_val
                st.rerun()

        st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    has_approved = any(v["approved"] for v in st.session_state.table_review_import.values())
    approved_export_map = {
        full_name: export_map[full_name]
        for full_name, data in st.session_state.table_review_import.items()
        if data["approved"] and full_name in export_map
    }

    st.divider()
    st.markdown(
        "<div class='section-title'>📋 Export as PySpark</div>"
        "<div class='section-caption'>One combined script with a COMMENT ON TABLE statement "
        "for every selected table.</div>",
        unsafe_allow_html=True,
    )
    py_multi = _build_multi_table_pyspark_script(export_map)
    exp1, exp2, _ = st.columns([1.8, 1.8, 4.4])
    exp1.download_button(
        "⬇ Download (All)",
        data=py_multi,
        file_name="table_descriptions_all.py",
        mime="text/plain",
        use_container_width=True,
    )
    if has_approved:
        py_multi_approved = _build_multi_table_pyspark_script(approved_export_map)
        exp2.download_button(
            "⬇ Download (Approved only)",
            data=py_multi_approved,
            file_name="table_descriptions_approved.py",
            mime="text/plain",
            use_container_width=True,
        )

    with st.expander("Preview PySpark (All)", expanded=False):
        st.code(py_multi, language="python")
    if has_approved:
        with st.expander("Preview PySpark (Approved only)", expanded=False):
            st.code(py_multi_approved, language="python")

    if st.button("🗑 Clear", use_container_width=False):
        for name in selected_names:
            tbl_meta = next((t for t in st.session_state.tables if t["qualified_name"] == name), None)
            if tbl_meta:
                st.session_state.pop(f"text_tabledesc_{tbl_meta['full_name']}", None)
        st.session_state.table_descriptions = {}
        st.session_state.table_gen_errors = {}
        st.session_state.table_review_import = {}
        st.session_state.generated = False
        st.session_state.gen_error = None
        st.rerun()

    st.stop()

tbl = st.session_state.table_meta
cols = st.session_state.columns

if not tbl:
    st.markdown("""
<div class='empty-state'>
  <div class='es-icon'>📖</div>
  <h3>Select a table to get started</h3>
  <p>Browse your Unity Catalog from the sidebar, then generate AI descriptions with one click.</p>
  <br>
  <div>
    <span class='step'><span class='step-num'>1</span>Connect to Databricks</span>
    <span class='step'><span class='step-num'>2</span>Pick a catalog → schema → table</span>
    <span class='step'><span class='step-num'>3</span>Click ⚡ Generate &amp; Humanize</span>
    <span class='step'><span class='step-num'>4</span>Review, export &amp; apply in Databricks</span>
  </div>
</div>
""", unsafe_allow_html=True)
    st.stop()

type_label = tbl.get("type_label", "Delta Table")
st.markdown(
    f"### `{tbl['full_name']}`{_type_badge_html(type_label)}",
    unsafe_allow_html=True,
)

# ── Table overview ────────────────────────────────────────────────────────
st.markdown("<div class='overview-label'>Table Overview</div>", unsafe_allow_html=True)
td_left, td_right = st.columns([6, 1])
table_desc = td_left.text_area(
    "table_overview",
    value=st.session_state.table_description or tbl.get("comment", ""),
    height=80,
    label_visibility="collapsed",
    key="text_table_desc",
    placeholder="Click ⚡ Generate & Humanize to produce an AI overview, or type one manually.",
)
st.session_state.table_description = table_desc
with td_right:
    st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
    if st.button("✨", key="hz_table_desc", help="Re-humanize table description"):
        with st.spinner("Humanizing…"):
            humanized_td = hz.humanize(table_desc)
        st.session_state.table_description = humanized_td
        st.session_state["text_table_desc"] = humanized_td
        st.rerun()
if tbl.get("comment"):
    st.caption(f"Current in Databricks: {tbl['comment']}")

if not cols:
    st.info("No columns found for this table, or you may not have SELECT access.")
    st.stop()

if not st.session_state.generated:
    st.markdown(
        "<div class='hint-card'>⚡ Click <strong>Generate &amp; Humanize</strong> in the sidebar "
        "to produce AI descriptions — or type directly in the suggestion fields below.</div>",
        unsafe_allow_html=True,
    )

st.divider()

# ── Share for Review / Import Reviewed File ───────────────────────────────

export_col, import_col = st.columns(2)

with export_col:
    st.markdown("<div class='section-title'>📤 Share for Review</div><div class='section-caption'>Download an Excel file for your team. They fill in the approval + notes columns and return it.</div>", unsafe_allow_html=True)
    if cols:
        xlsx = _build_excel(tbl["full_name"], cols, st.session_state.suggestions, st.session_state.table_description)
        st.download_button(
            label="Download Review File (.xlsx)",
            data=xlsx,
            file_name=f"{tbl['full_name'].replace('.', '_')}_review.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

with import_col:
    st.markdown("<div class='section-title'>📥 Import Reviewed File</div><div class='section-caption'>Upload the completed file (.xlsx or .csv). Approved rows load back into the grid automatically.</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Upload reviewed file", type=["xlsx", "csv"],
        label_visibility="collapsed", key="review_upload",
    )
    if uploaded:
        try:
            with st.spinner("Reading reviewed file…"):
                raw = uploaded.read()
                review_data = (
                    _parse_csv(raw) if uploaded.name.endswith(".csv")
                    else _parse_excel(raw)
                )
            approved_count = sum(1 for v in review_data.values() if v["approved"])
            rejected_count = len(review_data) - approved_count

            for col_name, data in review_data.items():
                if data["approved"] and data["proposed_description"]:
                    val = data["proposed_description"]
                    st.session_state.suggestions[col_name] = val
                    st.session_state[f"text_{col_name}"] = val  # keep widget in sync
            st.session_state.review_import = review_data
            st.session_state.apply_results = {}

            st.success(f"✅ {approved_count} approved  •  ⛔ {rejected_count} not approved")
            st.rerun()
        except Exception as e:
            st.error(f"Could not read file: {e}")

st.divider()

# ── Column grid ───────────────────────────────────────────────────────────

h1, h2, h3, h4, h5 = st.columns([2, 1.2, 3, 3, 1])
h1.markdown("<div class='col-label'>Column</div>", unsafe_allow_html=True)
h2.markdown("<div class='col-label'>Type</div>", unsafe_allow_html=True)
h3.markdown("<div class='col-label'>Current Description</div>", unsafe_allow_html=True)
h4.markdown("<div class='col-label'>AI Suggestion (editable)</div>", unsafe_allow_html=True)
h5.markdown("<div class='col-label'>✨</div>", unsafe_allow_html=True)
st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

for col in cols:
    name = col["name"]
    review = st.session_state.review_import.get(name)
    c1, c2, c3, c4, c5 = st.columns([2, 1.2, 3, 3, 1])

    review_html = ""
    if review:
        if review["approved"]:
            review_html += "<span class='chip chip-approved'>✓ Approved</span>"
        else:
            review_html += "<span class='chip chip-rejected'>✗ Rejected</span>"
        if review["notes"]:
            review_html += f"<br><span class='chip chip-note'>💬 {review['notes']}</span>"

    c1.markdown(
        f"<div class='col-name'>{name}</div>{review_html}",
        unsafe_allow_html=True,
    )

    c2.markdown(_type_pill(col["type"]), unsafe_allow_html=True)
    c3.markdown(
        f"<div class='current-desc'>{col['current_comment'] or '<span style=\"color:#2e3755\">—</span>'}</div>",
        unsafe_allow_html=True,
    )

    new_text = c4.text_area(
        label=f"suggestion_{name}",
        value=st.session_state.suggestions.get(name, col["current_comment"]),
        height=90, label_visibility="collapsed", key=f"text_{name}",
    )
    st.session_state.suggestions[name] = new_text

    with c5:
        st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
        if st.button("✨", key=f"hz_{name}", help=f"Re-humanize {name}"):
            with st.spinner("Humanizing…"):
                val = hz.humanize(st.session_state.suggestions.get(name, ""))
            st.session_state.suggestions[name] = val
            st.session_state[f"text_{name}"] = val  # keep widget in sync
            st.rerun()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)


# ── Action buttons ────────────────────────────────────────────────────────

st.markdown("")
has_approved = any(v["approved"] for v in st.session_state.review_import.values())
approved_only = {
    name: st.session_state.suggestions[name]
    for name, data in st.session_state.review_import.items()
    if data["approved"] and name in st.session_state.suggestions
}

if st.button("🗑 Clear", use_container_width=False):
    for col in st.session_state.columns:
        st.session_state.pop(f"text_{col['name']}", None)
    st.session_state.pop("text_table_desc", None)
    st.session_state.suggestions = {}
    st.session_state.table_description = ""
    st.session_state.review_import = {}
    st.session_state.generated = False
    st.session_state.gen_error = None
    st.rerun()

# ── Export as code ────────────────────────────────────────────────────────

st.markdown("<div class='section-title'>📋 Export as PySpark</div><div class='section-caption'>Download or copy the script — paste into a Databricks notebook to apply descriptions.</div>", unsafe_allow_html=True)

safe_name = tbl["full_name"].replace(".", "_")
table_desc_export = st.session_state.table_description
py_all = _build_pyspark_script(tbl["full_name"], st.session_state.suggestions, table_desc_export)

if has_approved:
    py_app = _build_pyspark_script(tbl["full_name"], approved_only, table_desc_export)

exp1, exp2, _ = st.columns([1.8, 1.8, 4.4])
exp1.download_button("⬇ Download (All)", data=py_all,
    file_name=f"{safe_name}_all.py", mime="text/plain", use_container_width=True)
if has_approved:
    exp2.download_button("⬇ Download (Approved only)", data=py_app,
        file_name=f"{safe_name}_approved.py", mime="text/plain", use_container_width=True)

with st.expander("Preview PySpark (All)", expanded=False):
    st.code(py_all, language="python")

if has_approved:
    with st.expander("Preview PySpark (Approved only)", expanded=False):
        st.code(py_app, language="python")
